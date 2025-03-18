import openpyxl

# Criar uma planilha
book = openpyxl.Workbook()

# Visualizar páginas existentes
print(book.sheetnames)

# Criar página, pedindo para colocar um nome
pagina = input("Qual nome da página? ")
book.create_sheet(pagina)

# Selecionar página
pagina_page = book[pagina]

# Perguntar quantas colunas adicionar
num_colunas = int(input("Quantas colunas deseja adicionar? "))

# Criar uma lista para armazenar nome das colunas
lista_coluna = []

# Loop para coletar nome das colunas
for i in range(num_colunas):
    nome_coluna = input(f"Informe o nome da coluna {i+1}: ")
    lista_coluna.append(nome_coluna)  # Adiciona os nomes na lista

# Adiciona os nomes das colunas na primeira linha da planilha
pagina_page.append(lista_coluna)

# Perguntar quantas linhas de dados o usuário deseja adicionar
num_linhas = int(input("Quantas linhas de dados deseja adicionar? "))

# Loop para coletar os valores das colunas e adicionar nas linhas
for i in range(num_linhas):
    linha_valores = []
    print(f"Informe os valores para a linha {i+1}:")
    for j in range(num_colunas):
        valor = input(f"Valor para a coluna {lista_coluna[j]}: ")
        linha_valores.append(valor)
    pagina_page.append(linha_valores)

# Salvar a planilha
book.save('Planilha de Frutas2.xlsx')
